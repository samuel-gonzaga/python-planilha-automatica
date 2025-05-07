from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# 1. Abrir arquivos
wb_master = load_workbook('master.xlsx')
wb_template = load_workbook('lista_chamada.xlsx')

# 2. Copiar dados
sheet_1 = wb_master['Pannunzio - Sab 8h-10h']
sheet_2 = wb_master['Pannunzio - Sab 10h-12h']
sheet_3 = wb_master['Pannunzio - Sab 12h30-14h30']
sheet_4 = wb_master['Pannunzio - Sab 14h30-16h30']

sheet_output = wb_template['Table 1']

# 3. Pegar nomes (coluna C, a partir da linha 3)
sala_6_1 = [sheet_1.cell(row=i, column=1).value for i in range(5, 51) if sheet_1.cell(row=i, column=1).value]
sala_7_1 = [sheet_1.cell(row=i, column=1).value for i in range(55, 101) if sheet_1.cell(row=i, column=1).value]

sala_6_2 = [sheet_2.cell(row=i, column=1).value for i in range(5, 51) if sheet_2.cell(row=i, column=1).value]
sala_7_2 = [sheet_2.cell(row=i, column=1).value for i in range(55, 101) if sheet_2.cell(row=i, column=1).value]

sala_6_3 = [sheet_3.cell(row=i, column=1).value for i in range(5, 51) if sheet_3.cell(row=i, column=1).value]
sala_7_3 = [sheet_3.cell(row=i, column=1).value for i in range(55, 101) if sheet_3.cell(row=i, column=1).value]

sala_6_4 = [sheet_4.cell(row=i, column=1).value for i in range(5, 51) if sheet_4.cell(row=i, column=1).value]
sala_7_4 = [sheet_4.cell(row=i, column=1).value for i in range(55, 101) if sheet_4.cell(row=i, column=1).value]

# 4. Colar na lista (coluna H, a partir da linha 9)

# 08:00 ATÉ 10:00 - SALA 06
for i, nome in enumerate(sala_6_1, start=5):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome

# 08:00 ATÉ 10:00 - SALA 07
for i, nome in enumerate(sala_7_1, start=66):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome

# 10:00 até 12:00 - SALA 06
for i, nome in enumerate(sala_6_2, start=129):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome

# 10:00 até 12:00 - SALA 07
for i, nome in enumerate(sala_7_2, start=188):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome

# 12:30 até 14:30 - SALA 06
for i, nome in enumerate(sala_6_3, start=248):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome

# 12:30 até 14:30 - SALA 07
for i, nome in enumerate(sala_7_3, start=290):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome

# 14:30 até 16:30 - SALA 06
for i, nome in enumerate(sala_6_4, start=336):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome

# 14:30 até 16:30 - SALA 07
for i, nome in enumerate(sala_7_4, start=394):
    cell = sheet_output.cell(row=i, column=8)
    if not isinstance(cell, MergedCell):
        cell.value = nome


# 5. Salvar
wb_template.save('lista_chamada_atualizada.xlsx')