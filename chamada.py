from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# Mapeia cada aula com: (planilha, faixa de linhas da sala 6, faixa da sala 7, linha inicial de escrita para cada sala)
CONFIG = [
    ("Pannunzio - Sab 8h-10h", (5, 50), (55, 105), 5, 66),
    ("Pannunzio - Sab 10h-12h", (5, 50), (55, 105), 129, 188),
    ("Pannunzio - Sab 12h30-14h30", (5, 50), (55, 105), 248, 290),
    ("Pannunzio - Sab 14h30-16h30", (5, 50), (55, 105), 336, 394),
]

def extrair_nomes(sheet, start_row, end_row):
    """Extrai os nomes da coluna C (3) em uma faixa de linhas."""
    return [
        sheet.cell(row=i, column=3).value
        for i in range(start_row, end_row)
        if sheet.cell(row=i, column=3).value
    ]

def colar_nomes(sheet_destino, nomes, start_row):
    """Cola os nomes na coluna H (8) a partir de uma linha inicial."""
    for i, nome in enumerate(nomes, start=start_row):
        cell = sheet_destino.cell(row=i, column=8)
        if not isinstance(cell, MergedCell):
            cell.value = nome

def processar_chamada(caminho_master, caminho_template, caminho_saida):
    wb_master = load_workbook(caminho_master)
    wb_template = load_workbook(caminho_template)
    sheet_output = wb_template['Table 1']

    for config in CONFIG:
        aba, faixa_sala6, faixa_sala7, linha_saida_6, linha_saida_7 = config
        sheet = wb_master[aba]

        nomes_sala6 = extrair_nomes(sheet, *faixa_sala6)
        nomes_sala7 = extrair_nomes(sheet, *faixa_sala7)

        colar_nomes(sheet_output, nomes_sala6, linha_saida_6)
        colar_nomes(sheet_output, nomes_sala7, linha_saida_7)

    wb_template.save(caminho_saida)
